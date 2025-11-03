from __future__ import annotations
"""
Servicio de consolidación de deudas ART
---------------------------------------
Reglas principales:
• Cruce por **(CUIT + Aseguradora de ORIGEN)**. En el maestro se elige la fila de esa aseguradora con «Cuenta Perdida» = vacío o “Vigente”. Si no hay vigente, se usa la primera no vigente. Si no existe ninguna fila para ese par → va a «No cruzan».
• La columna «Aseguradora» del Consolidado sale del ORIGEN (archivo de deuda), aplicando alias (p.ej., OMINT → Serena).
• «Premier» sale de «Referido por (Nombre de Cuenta)» (PREMIER→Premier; otro→No es Premier).
• «Estado contrato» sale de «Cuenta Perdida» (vacío/“Vigente”→Vigente; otro→ese texto).
• «Productor» vacío → PROMECOR (en Consolidado).
• Q = Deuda/Costo si Costo=0 o vacío → Costo y Q vacíos.
• Se excluye de «Consolidado»: Deuda total entre 0 y 999 (incl.), y Ramo="Domestica".
• **Andina ART y Parana Seguros**: agrupar por CUIT y sumar (archivos tipo tabla dinámica).
• “Experta”: deuda con signo invertido → se invierte.
• Moneda ARS "$  #,##0.00"; CUIT 11 dígitos; zoom 80% en todas las hojas.

Incluye: parser robusto AR para importes + tolerancia a espacios en encabezados (strip) + normalización de acentos (slug) + alias OMINT→Serena.
"""

from io import BytesIO
from pathlib import Path
from typing import List, Dict, Set
import re
import unicodedata

import pandas as pd
from django.conf import settings


# ------------------------------------------------------------------#
# Rutas y columnas
# ------------------------------------------------------------------#
BASE_ASEG_DIR: Path = getattr(
    settings, "ART_ASEG_DIR",
    Path(r"C:/Users/Promecor/Documents/ART/Aseguradoras"),
)

MAESTRO_PATH: Path = getattr(
    settings, "ART_MAESTRO_PATH",
    Path(r"C:/Users/Promecor/Documents/Aplicativo cobranzas/ART/TOTAL PARA SUBIR DEUDAS ART.xlsx"),
)

MAPEO_ASEG_PATH: Path = getattr(
    settings, "ART_MAPEO_ASEG_PATH",
    Path(r"C:/Users/Promecor/Documents/Aplicativo cobranzas/ART/Mapeo aseguradoras.xlsx"),
)

COLUMNS_ORDER: List[str] = [
    "Periodo", "Razón social", "CUIT", "Contrato", "Aseguradora",
    "Deuda total", "Costo mensual", "Q periodos deudores", "Estado contrato",
    "Email del trato", "No contactar", "Productor", "Premier",
    "Cliente importante",
]

# Columnas EXACTAS del maestro:
M_CUIT           = "CUIT (Nombre de Cuenta)"
M_RAZON          = "Nombre de Cuenta (Nombre de Cuenta)"
M_CONTRATO       = "Número de contrato"
M_ASEGURADORA    = "Aseguradora Enviada LookUp"
M_COSTO          = "Aporte LRT (Nombre de Cuenta)"
M_CUENTA_PERDIDA = "Cuenta Perdida"
M_EMAIL          = "Correo electrónico"
M_NO_CONTACTAR   = "No Contactar"
M_PRODUCTOR1     = "Productor"
M_PRODUCTOR2     = "Productor (Nombre de Cuenta)"
M_REFERIDO_POR   = "Referido por (Nombre de Cuenta)"
M_CLIENTE_IMP    = "Cliente Importante (Nombre de Cuenta)"
M_CAPITAS        = "Cápitas (Nombre de Cuenta)"
M_RAMO           = "Ramo"

CAPITAS_COL_OUT = "Capitas"


# ------------------------------------------------------------------#
# Helpers
# ------------------------------------------------------------------#
def _as_bool(s: pd.Series) -> pd.Series:
    truthy = {"true", "verdadero", "1", "t", "yes", "si", "sí"}
    return s.astype(str).str.strip().str.casefold().isin(truthy)

def _is_empty_email(s: pd.Series) -> pd.Series:
    ss = s.astype(str).str.strip()
    return ss.eq("") | ss.str.lower().isin({"nan", "none"})

def _norm_cuit_str(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
         .str.replace(r"[^0-9]", "", regex=True)
         .str.zfill(11)
         .str[-11:]
    )

def _norm_aseg_str(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().str.upper()

def _slug(s: str) -> str:
    if s is None:
        return ""
    base = unicodedata.normalize("NFKD", str(s).strip().casefold())
    return "".join(ch for ch in base if not unicodedata.combining(ch))

def _alias_aseg_for_merge(name: str) -> str:
    """Alias de nombre de carpeta/origen a nombre de merge contra maestro."""
    key = _slug(name)
    alias_map = {
        "omint": "Serena",
        "omint art": "Serena",
    }
    return alias_map.get(key, name)

def _ensure_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = pd.NA
    return out[cols]

def _norm_periodo(periodo: str) -> str:
    p = (periodo or "").strip().replace("/", "-")
    mm, yyyy = p.split("-")
    return f"{mm.zfill(2)}-{yyyy}"

_AR_NUM_RE = re.compile(r"[^0-9,\.\-]")

def _to_number_ar_series(s: pd.Series, decimals: int | None = None) -> pd.Series:
    if pd.api.types.is_numeric_dtype(s):
        out = pd.to_numeric(s, errors="coerce")
    else:
        tmp = s.astype(str).str.strip().replace({"": None, "nan": None, "None": None})
        tmp = tmp.apply(lambda x: None if x is None else _AR_NUM_RE.sub("", x))
        def _swap_commas(v):
            if v is None:
                return None
            if "," in v:
                v = v.replace(".", "").replace(",", ".")
            return v
        tmp = tmp.apply(_swap_commas)
        out = pd.to_numeric(tmp, errors="coerce")
    if decimals is not None:
        out = out.round(decimals)
    return out


# ------------------------------------------------------------------#
# Lectura de insumos
# ------------------------------------------------------------------#
def _leer_mapeo_aseguradoras(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0, dtype=str)
    df.columns = df.columns.str.strip()
    need = {"Aseguradora", "deuda_col", "cuit_col"}
    faltan = need - set(df.columns)
    if faltan:
        raise ValueError(f"Mapeo aseguradoras incompleto. Faltan columnas: {faltan}")
    for col in ["Aseguradora", "deuda_col", "cuit_col"]:
        df[col] = df[col].astype(str).str.strip()
    return df

def _cargar_maestro_raw(path: Path) -> pd.DataFrame:
    use_cols = [
        M_CUIT, M_RAZON, M_CONTRATO, M_ASEGURADORA, M_COSTO,
        M_CUENTA_PERDIDA, M_EMAIL, M_NO_CONTACTAR, M_PRODUCTOR1, M_PRODUCTOR2,
        M_REFERIDO_POR, M_CLIENTE_IMP, M_CAPITAS, M_RAMO,
    ]
    maestro = pd.read_excel(path, sheet_name=0, dtype=str)
    presentes = [c for c in use_cols if c in maestro.columns]
    maestro = maestro[presentes].copy()
    for c in use_cols:
        if c not in maestro.columns:
            maestro[c] = pd.NA

    maestro[M_CUIT]  = _norm_cuit_str(maestro[M_CUIT])
    maestro[M_COSTO] = _to_number_ar_series(maestro[M_COSTO], decimals=2)
    return maestro

def _leer_deudas_archivo(fp: Path, nombre_aseg: str, mapeo: pd.DataFrame) -> pd.DataFrame:
    """
    Devuelve: ['cuit', 'deuda_total', 'aseguradora_origen']
      - Federación Patronal: permite deuda_col "X + Y".
      - Andina ART y Parana Seguros: agrupa por CUIT y suma.
      - Experta: invierte signo.
      - Alias de aseguradora para merge (p.ej., OMINT→Serena).
    """
    df = pd.read_excel(fp, sheet_name=0)
    df.columns = df.columns.str.strip()

    spec = mapeo[mapeo["Aseguradora"].astype(str).str.strip().str.casefold()
                 == nombre_aseg.strip().casefold()]
    if spec.empty:
        raise ValueError(f"No hay mapeo para '{nombre_aseg}'. Verificá 'Mapeo aseguradoras.xlsx'.")

    cuit_col = str(spec.iloc[0]["cuit_col"]).strip()
    deuda_col = str(spec.iloc[0]["deuda_col"]).strip()

    def _check_cols(cols: List[str]):
        miss = [c for c in cols if c not in df.columns]
        if miss:
            raise KeyError(
                f"En '{nombre_aseg}' faltan columnas {miss} en {fp.name}. "
                f"Encabezados disponibles: {list(df.columns)}"
            )

    if isinstance(deuda_col, str) and "+" in deuda_col:
        partes = [p.strip() for p in deuda_col.split("+")]
        _check_cols(partes + [cuit_col])
        tot = None
        for p in partes:
            serie = _to_number_ar_series(df[p], decimals=2)
            tot = serie if tot is None else tot.add(serie, fill_value=0)
        deuda_series = tot
    else:
        _check_cols([deuda_col, cuit_col])
        deuda_series = _to_number_ar_series(df[deuda_col], decimals=2)

    if "experta" in _slug(nombre_aseg):
        deuda_series = deuda_series * -1

    aseg_for_merge = _alias_aseg_for_merge(nombre_aseg)

    tmp = pd.DataFrame({
        "cuit": _norm_cuit_str(df[cuit_col]),
        "deuda_total": deuda_series,
        "aseguradora_origen": aseg_for_merge,
    })

    aseg_key = _slug(nombre_aseg)
    if ("andina" in aseg_key) or ("parana" in aseg_key):
        tmp = (
            tmp.groupby("cuit", as_index=False, sort=False)["deuda_total"]
               .sum()
               .assign(aseguradora_origen=aseg_for_merge)
        )

    return tmp[["cuit", "deuda_total", "aseguradora_origen"]]

def _cargar_deudas(periodo: str, mapeo: pd.DataFrame) -> pd.DataFrame:
    fn = f"{_norm_periodo(periodo)}.xlsx"
    dfs: List[pd.DataFrame] = []

    base = BASE_ASEG_DIR
    if not base.exists():
        raise FileNotFoundError(f"No existe la carpeta de Aseguradoras: {base}")

    for carpeta in sorted([p for p in base.iterdir() if p.is_dir()]):
        fp = carpeta / fn
        if not fp.exists():
            continue
        dfi = _leer_deudas_archivo(fp, carpeta.name, mapeo)
        dfs.append(dfi)

    if not dfs:
        raise FileNotFoundError(f"No hay archivos {fn} en {base}")

    deudas = pd.concat(dfs, ignore_index=True)
    deudas = (
        deudas.groupby(["cuit", "aseguradora_origen"], as_index=False, sort=False)["deuda_total"]
              .sum()
    )
    return deudas


# ------------------------------------------------------------------#
# Hoja «Consolidado»
# ------------------------------------------------------------------#
def df_consolidado(periodo: str) -> pd.DataFrame:
    maestro = _cargar_maestro_raw(MAESTRO_PATH)
    mapeo = _leer_mapeo_aseguradoras(MAPEO_ASEG_PATH)
    deudas = _cargar_deudas(periodo, mapeo)

    # Vigente: vacío o literal "Vigente"
    cp = maestro[M_CUENTA_PERDIDA].astype(str).str.strip().str.casefold()
    is_vigente = maestro[M_CUENTA_PERDIDA].isna() | cp.isin({"", "vigente"})
    maestro["Vigente"] = is_vigente

    maestro["_ASEG_n"] = _norm_aseg_str(maestro[M_ASEGURADORA])
    deudas["_ASEG_ORIGEN_n"] = _norm_aseg_str(deudas["aseguradora_origen"])

    maestro_sorted = maestro.sort_values([M_CUIT, "_ASEG_n", "Vigente"], ascending=[True, True, False])
    maestro_compacto = maestro_sorted.drop_duplicates([M_CUIT, "_ASEG_n"], keep="first")

    df = deudas.merge(
        maestro_compacto,
        how="left",
        left_on=["cuit", "_ASEG_ORIGEN_n"],
        right_on=[M_CUIT, "_ASEG_n"],
        suffixes=("", "_m"),
    )

    df = df[df["_ASEG_n"].notna()].copy()

    df["Periodo"]          = _norm_periodo(periodo)
    df["Razón social"]     = df[M_RAZON]
    df["CUIT"]             = df[M_CUIT]
    df["Contrato"]         = pd.to_numeric(df[M_CONTRATO], errors="coerce")
    df["Aseguradora"]      = df["aseguradora_origen"]
    df["Deuda total"]      = _to_number_ar_series(df["deuda_total"], decimals=2)
    df["Costo mensual"]    = _to_number_ar_series(df[M_COSTO], decimals=2)

    df["Q periodos deudores"] = df.apply(
        lambda r: round(r["Deuda total"] / r["Costo mensual"], 2)
        if pd.notna(r["Costo mensual"]) and r["Costo mensual"] else None,
        axis=1,
    )
    df.loc[(df["Costo mensual"].isna()) | (df["Costo mensual"].eq(0)), ["Costo mensual", "Q periodos deudores"]] = pd.NA

    df["Estado contrato"] = df[M_CUENTA_PERDIDA].apply(
        lambda x: "Vigente" if (pd.isna(x) or str(x).strip().casefold() in {"", "vigente"}) else x
    )

    df["Premier"] = df[M_REFERIDO_POR].apply(
        lambda x: "Premier" if str(x).strip().upper() == "PREMIER" else "No es Premier"
    )

    df["Email del trato"]  = df[M_EMAIL]
    df["No contactar"]     = df[M_NO_CONTACTAR]

    if M_PRODUCTOR1 in df.columns and df[M_PRODUCTOR1].notna().any():
        prod = df[M_PRODUCTOR1]
    else:
        prod = df[M_PRODUCTOR2]
    df["Productor"] = prod.where(prod.astype(str).str.strip() != "", "PROMECOR")

    df["Cliente importante"] = df[M_CLIENTE_IMP]

    ramo_norm = df.get(M_RAMO, pd.Series("", index=df.index)).astype(str).str.strip().str.casefold()
    df = df[~ramo_norm.eq("domestica")]

    deuda_num = _to_number_ar_series(df["Deuda total"])
    df = df[(deuda_num.ge(1000)) | (deuda_num.lt(0))]

    out = df[COLUMNS_ORDER].copy()
    out.attrs["cuits_ambig"] = set()
    out.attrs["cuits_con_vigente_unico"] = set()
    return out


# ------------------------------------------------------------------#
# Hoja «No cruzan»
# ------------------------------------------------------------------#
def df_no_cruzan(periodo: str, cuits_duplicados: Set[str]) -> pd.DataFrame:
    mapeo = _leer_mapeo_aseguradoras(MAPEO_ASEG_PATH)
    deudas = _cargar_deudas(periodo, mapeo)

    maestro = _cargar_maestro_raw(MAESTRO_PATH)
    cp = maestro[M_CUENTA_PERDIDA].astype(str).str.strip().str.casefold()
    is_vigente = maestro[M_CUENTA_PERDIDA].isna() | cp.isin({"", "vigente"})
    maestro["Vigente"] = is_vigente

    maestro["_ASEG_n"] = _norm_aseg_str(maestro[M_ASEGURADORA])
    deudas["_ASEG_ORIGEN_n"] = _norm_aseg_str(deudas["aseguradora_origen"])

    maestro_sorted = maestro.sort_values([M_CUIT, "_ASEG_n", "Vigente"], ascending=[True, True, False])
    maestro_compacto = maestro_sorted.drop_duplicates([M_CUIT, "_ASEG_n"], keep="first")

    merged = deudas.merge(
        maestro_compacto,
        how="left",
        left_on=["cuit", "_ASEG_ORIGEN_n"],
        right_on=[M_CUIT, "_ASEG_n"],
        suffixes=("", "_m"),
    )

    df_nc = merged[merged["_ASEG_n"].isna()].copy()

    df_nc["Periodo"]             = _norm_periodo(periodo)
    df_nc["Razón social"]        = pd.NA
    df_nc["CUIT"]                = df_nc["cuit"]
    df_nc["Contrato"]            = pd.NA
    df_nc["Aseguradora"]         = df_nc["aseguradora_origen"]
    df_nc["Deuda total"]         = _to_number_ar_series(df_nc["deuda_total"], decimals=2)
    df_nc["Costo mensual"]       = pd.NA
    df_nc["Q periodos deudores"] = pd.NA
    df_nc["Estado contrato"]     = pd.NA
    df_nc["Email del trato"]     = pd.NA
    df_nc["No contactar"]        = pd.NA
    df_nc["Productor"]           = pd.NA
    df_nc["Premier"]             = pd.NA
    df_nc["Cliente importante"]  = pd.NA

    out = df_nc[COLUMNS_ORDER].copy()
    return out


# ------------------------------------------------------------------#
# Listados basados en «Consolidado»
# ------------------------------------------------------------------#
def df_sin_mail(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    return base[email_vacio & (base["Premier"] == "No es Premier")]

def df_anuladas(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    return base[base["Estado contrato"].astype(str).str.strip().ne("Vigente") & (~email_vacio)]

def df_no_contactar(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    return base[
        _as_bool(base["No contactar"]) &
        (~_as_bool(base["Cliente importante"])) &
        base["Estado contrato"].eq("Vigente") &
        (~email_vacio)
    ]

def df_clientes_importantes(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    return base[
        _as_bool(base["Cliente importante"]) &
        (~_as_bool(base["No contactar"])) &
        base["Estado contrato"].eq("Vigente") &
        (~email_vacio)
    ]

def df_un_q_deudor(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    q = pd.to_numeric(base["Q periodos deudores"], errors="coerce")
    return base[
        q.notna() & (q <= 1) &
        base["Estado contrato"].eq("Vigente") &
        (~email_vacio)
    ]

def df_premier(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    return base[(base["Premier"] == "Premier") & base["Estado contrato"].eq("Vigente") & (~email_vacio)]


# ------------------------------------------------------------------#
# Más listados derivados
# ------------------------------------------------------------------#
def df_productor(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    q = pd.to_numeric(base["Q periodos deudores"], errors="coerce")
    deuda = pd.to_numeric(base["Deuda total"], errors="coerce")
    return base[
        (base["Productor"].fillna("").str.strip().str.upper() != "PROMECOR") &
        q.notna() & (q > 1) &
        base["Estado contrato"].eq("Vigente") &
        (~email_vacio) &
        deuda.ge(1000)
    ]

def df_deuda_promecor(base: pd.DataFrame) -> pd.DataFrame:
    email_vacio = _is_empty_email(base["Email del trato"])
    q = pd.to_numeric(base["Q periodos deudores"], errors="coerce")
    deuda = pd.to_numeric(base["Deuda total"], errors="coerce")
    return base[
        (base["Productor"].fillna("").str.strip().str.upper() == "PROMECOR") &
        q.notna() & (q > 1) &
        base["Estado contrato"].eq("Vigente") &
        (~_as_bool(base["Cliente importante"])) &
        (~_as_bool(base["No contactar"])) &
        (base["Premier"] == "No es Premier") &
        (~email_vacio) &
        deuda.ge(1000)
    ]


# ------------------------------------------------------------------#
# «Agregar costo mensual»
# ------------------------------------------------------------------#
def df_agregar_costo_mensual(base: pd.DataFrame) -> pd.DataFrame:
    maestro = _cargar_maestro_raw(MAESTRO_PATH)
    maestro_cap = maestro[[M_CUIT, M_CAPITAS]].rename(columns={M_CUIT: "CUIT"})

    aux = base.merge(maestro_cap, on="CUIT", how="left")
    aux["Costo mensual"] = _to_number_ar_series(aux["Costo mensual"], decimals=2)

    mask = aux["Costo mensual"].isna() | aux["Costo mensual"].eq(0)
    out = aux.loc[mask].copy()
    out["Q periodos deudores"] = pd.NA

    out[CAPITAS_COL_OUT] = pd.to_numeric(out[M_CAPITAS], errors="coerce")

    cols = COLUMNS_ORDER + [CAPITAS_COL_OUT]
    return _ensure_columns(out, cols)


# ------------------------------------------------------------------#
# Exportador genérico
# ------------------------------------------------------------------#
def _exportar_excel(hojas: Dict[str, pd.DataFrame], destino: Path | BytesIO) -> None:
    import xlsxwriter
    from openpyxl import load_workbook

    MONEY_FMT = '"$ " #,##0.00'
    Q_FMT     = '0.00'
    CUIT_FMT  = "00000000000"
    INT_FMT   = "0"

    tmp = BytesIO()
    with pd.ExcelWriter(tmp, engine="xlsxwriter") as writer:
        for nombre, df in hojas.items():
            safe = df.copy()

            for c in ["Deuda total", "Costo mensual", "Q periodos deudores", "Contrato", "Capitas"]:
                if c in safe.columns:
                    safe[c] = pd.to_numeric(safe[c], errors="coerce")

            safe.to_excel(writer, sheet_name=nombre, index=False)
            wb, ws = writer.book, writer.sheets[nombre]

            fmt_cuit  = wb.add_format({"num_format": "0"})
            fmt_money = wb.add_format({"num_format": MONEY_FMT, "align": "right"})
            fmt_q     = wb.add_format({"num_format": Q_FMT, "align": "right"})
            fmt_int   = wb.add_format({"num_format": INT_FMT, "align": "right"})

            cols_meta = []
            for col in safe.columns:
                if col.upper().startswith("CUIT"):
                    cols_meta.append({"header": col, "format": fmt_cuit})
                elif col in ("Deuda total", "Costo mensual"):
                    cols_meta.append({"header": col, "format": fmt_money})
                elif col == "Q periodos deudores":
                    cols_meta.append({"header": col, "format": fmt_q})
                elif col in ("Contrato", "Capitas"):
                    cols_meta.append({"header": col, "format": fmt_int})
                else:
                    cols_meta.append({"header": col})

            r, c = safe.shape
            ws.add_table(0, 0, max(r, 1), max(c - 1, 0), {"style": "Table Style Medium 2", "columns": cols_meta})

            ancho = {
                "Periodo": 10, "Razón social": 35, "CUIT": 14, "Contrato": 14, "Aseguradora": 18,
                "Deuda total": 16, "Costo mensual": 16, "Q periodos deudores": 14, "Estado contrato": 18,
                "Email del trato": 34, "No contactar": 12, "Productor": 14, "Premier": 14,
                "Cliente importante": 16, "Capitas": 14,
            }
            for i, col in enumerate(safe.columns):
                ws.set_column(i, i, ancho.get(col, 12))
            ws.freeze_panes(1, 0)
            ws.set_zoom(80)

    tmp.seek(0)
    wb2 = load_workbook(tmp)

    def _set_col_format(ws, header_name: str, fmt: str, transform=None):
        headers = [cell.value for cell in ws[1]]
        if header_name in headers:
            j = headers.index(header_name) + 1
            col_letter = ws.cell(row=1, column=j).column_letter
            for cell in ws[col_letter][1:]:
                if transform:
                    try:
                        cell.value = transform(cell.value)
                    except Exception:
                        pass
                cell.number_format = fmt

    for nombre, df in hojas.items():
        ws2 = wb2[nombre]
        _set_col_format(ws2, "CUIT", CUIT_FMT, transform=lambda v: int(str(v).strip()) if str(v).strip() not in ("", "None", "nan") else None)
        _set_col_format(ws2, "Deuda total", MONEY_FMT)
        _set_col_format(ws2, "Costo mensual", MONEY_FMT)
        _set_col_format(ws2, "Q periodos deudores", Q_FMT)
        _set_col_format(ws2, "Contrato", INT_FMT)
        _set_col_format(ws2, "Capitas", INT_FMT)

    if isinstance(destino, BytesIO):
        destino.seek(0)
        wb2.save(destino)
        destino.seek(0)
    else:
        wb2.save(destino)


# ------------------------------------------------------------------#
# API pública
# ------------------------------------------------------------------#
def generar_xlsx(periodo: str) -> BytesIO:
    base = df_consolidado(periodo)
    cuits_dup = base.attrs.get("cuits_ambig", set())

    hojas: Dict[str, pd.DataFrame] = {
        "Consolidado":           _ensure_columns(base, COLUMNS_ORDER),
        "No cruzan":             _ensure_columns(df_no_cruzan(periodo, cuits_dup), COLUMNS_ORDER),
        "Sin mail":              _ensure_columns(df_sin_mail(base), COLUMNS_ORDER),
        "Anuladas":              _ensure_columns(df_anuladas(base), COLUMNS_ORDER),
        "No contactar":          _ensure_columns(df_no_contactar(base), COLUMNS_ORDER),
        "Clientes importantes":  _ensure_columns(df_clientes_importantes(base), COLUMNS_ORDER),
        "1 Q.deudor":            _ensure_columns(df_un_q_deudor(base), COLUMNS_ORDER),
        "Premier":               _ensure_columns(df_premier(base), COLUMNS_ORDER),
        "Productor":             _ensure_columns(df_productor(base), COLUMNS_ORDER),
        "Deuda Promecor":        _ensure_columns(df_deuda_promecor(base), COLUMNS_ORDER),
        "Agregar costo mensual": df_agregar_costo_mensual(base),
    }

    buf = BytesIO()
    _exportar_excel(hojas, buf)
    buf.seek(0)
    return buf
