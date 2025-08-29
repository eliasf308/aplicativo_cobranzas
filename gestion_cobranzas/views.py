"""gestion_cobranzas.views
    Archivo completo con rutas de templates actualizadas al nuevo
    esquema de carpetas. 05‑AUG‑2025
"""

from datetime import datetime, date
from decimal import Decimal
from functools import singledispatch
from pathlib import Path
import io
import re
import unicodedata

import numpy as np
import openpyxl
import pandas as pd
import weasyprint
import xlsxwriter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template, render_to_string
from django.utils import timezone
from django.utils.encoding import smart_str
from openpyxl.utils import get_column_letter
from weasyprint import CSS, HTML

from art.forms import (CargaMasivaForm, CuotaFormSet, ImputacionExcelForm,
                       PlanPagoForm)
# --- Alias provisorio mientras terminamos la mudanza de ART ---
from art.views.enviar_mails import enviar_mails_art  # noqa: F401
from .models import (Aseguradora, Cuota, LogCargaMasiva, LogImputacion, PlanPago,
                     Poliza, Ramo)
from .parsers.art_parsers import (_clean_number, parse_andina, parse_asociart,
                                  parse_berkley, parse_experta, parse_fede_patr,
                                  parse_galeno, parse_la_segunda, parse_omint,
                                  parse_prevencion, parse_provincia,
                                  parse_smg)

# ──────────────────── Carga masiva de planes ───────────────────────────

@login_required
def cargar_planes_excel(request):
    """Carga de un Excel con planes nuevos."""
    if request.method == "POST":
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            df = pd.read_excel(archivo)

            columnas_requeridas = [
                "Aseguradora",
                "Ramo",
                "Poliza",
                "Endoso",
                "Cuota",
                "Vencimiento",
                "Importe",
                "Moneda",
            ]
            if not all(col in df.columns for col in columnas_requeridas):
                messages.error(
                    request,
                    "El archivo debe contener las columnas: "
                    + ", ".join(columnas_requeridas),
                )
                return redirect("cargar_planes_excel")

            resumen = []
            grupos = df.groupby(["Aseguradora", "Ramo", "Poliza", "Endoso", "Moneda"])

            for (aseg, ramo, pol, endoso, moneda), grupo in grupos:
                aseguradora_obj, _ = Aseguradora.objects.get_or_create(nombre=aseg)
                ramo_obj, _ = Ramo.objects.get_or_create(nombre=ramo)
                poliza_obj, _ = Poliza.objects.get_or_create(
                    numero=str(pol), aseguradora=aseguradora_obj, ramo=ramo_obj
                )

                existe = PlanPago.objects.filter(
                    aseguradora=aseguradora_obj,
                    ramo=ramo_obj,
                    poliza=poliza_obj,
                    endoso=str(endoso),
                ).exists()

                if existe:
                    messages.warning(
                        request,
                        f"Ya existe un plan para {aseg} - {ramo} - {pol} - Endoso {endoso}. Se omitió la carga.",
                    )
                    continue

                plan = PlanPago.objects.create(
                    aseguradora=aseguradora_obj,
                    ramo=ramo_obj,
                    poliza=poliza_obj,
                    endoso=str(endoso),
                    moneda=moneda,
                )

                cuotas = []
                for _, row in grupo.iterrows():
                    importe = Decimal(row["Importe"])
                    cuota = Cuota(
                        plan_pago=plan,
                        numero=int(row["Cuota"]),
                        vencimiento=pd.to_datetime(row["Vencimiento"]).date(),
                        importe=importe,
                        importe_original=importe,
                    )
                    cuotas.append(cuota)
                Cuota.objects.bulk_create(cuotas)

                LogCargaMasiva.objects.create(
                    usuario=request.user,
                    aseguradora=aseg,
                    ramo=ramo,
                    poliza=str(pol),
                    endoso=str(endoso),
                    cantidad_cuotas=len(cuotas),
                    archivo=archivo.name,
                )

                resumen.append(
                    f"{aseg} - {ramo} - {pol} - Endoso {endoso}: {len(cuotas)} cuotas"
                )

            if resumen:
                messages.success(request, "Carga exitosa:\n" + "\n".join(resumen))
            else:
                messages.warning(request, "No se cargó ningún plan nuevo.")

            return redirect("cargar_planes_excel")
    else:
        form = CargaMasivaForm()

    return render(request, "gestion_cobranzas/planes/carga_excel.html", {"form": form})


# ────────────────────── Listado y detalle de planes ─────────────────────

@login_required
def listar_planes(request):
    poliza_query = request.GET.get("poliza", "")
    planes = PlanPago.objects.all()

    if poliza_query:
        planes = planes.filter(poliza__numero__icontains=poliza_query)

    total_saldo = Decimal("0.00")
    for plan in planes:
        plan.saldo = plan.cuotas.aggregate(restante=Sum("importe"))["restante"] or Decimal("0.00")
        total_saldo += plan.saldo

    return render(
        request,
        "gestion_cobranzas/planes/listado_planes.html",
        {
            "planes": planes,
            "poliza_query": poliza_query,
            "total_saldo": total_saldo,
        },
    )


@login_required
def ver_cuotas_plan(request, plan_id):
    plan = get_object_or_404(PlanPago, id=plan_id)
    cuotas = plan.cuotas.all().order_by("numero")

    pagadas = parciales = impagas = 0
    saldo_pendiente = total_premio = total_saldo = Decimal("0.00")

    for cuota in cuotas:
        saldo = cuota.importe
        premio = cuota.importe_original or Decimal("0.00")

        total_premio += premio
        total_saldo += saldo

        if saldo == 0:
            pagadas += 1
        elif saldo < premio:
            parciales += 1
        else:
            impagas += 1

        saldo_pendiente += saldo

    context = {
        "plan": plan,
        "cuotas": cuotas,
        "pagadas": pagadas,
        "parciales": parciales,
        "impagas": impagas,
        "saldo_pendiente": saldo_pendiente,
        "total_premio": total_premio,
        "total_saldo": total_saldo,
    }

    return render(request, "gestion_cobranzas/planes/cuotas_por_plan.html", context)


# ─────────────────────────── Edición de planes ──────────────────────────

@login_required
def eliminar_plan(request, plan_id):
    plan = get_object_or_404(PlanPago, id=plan_id)
    plan.delete()
    messages.success(request, "El plan fue eliminado correctamente.")
    return redirect("listar_planes")


@login_required
def editar_plan(request, plan_id):
    plan = get_object_or_404(PlanPago, id=plan_id)

    if request.method == "POST":
        form = PlanPagoForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, "El plan fue actualizado correctamente.")
            return redirect("listar_planes")
    else:
        form = PlanPagoForm(instance=plan)

    return render(request, "gestion_cobranzas/planes/editar_plan.html", {"form": form, "plan": plan})


@login_required
def editar_cuotas_plan(request, plan_id):
    plan = get_object_or_404(PlanPago, id=plan_id)
    cuotas = Cuota.objects.filter(plan_pago=plan).order_by("numero")

    if request.method == "POST":
        formset = CuotaFormSet(request.POST, queryset=cuotas)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Las cuotas fueron actualizadas correctamente.")
            return redirect("listar_planes")
    else:
        formset = CuotaFormSet(queryset=cuotas)

    return render(
        request,
        "gestion_cobranzas/planes/editar_cuotas_plan.html",
        {"plan": plan, "formset": formset},
    )


# ──────────────────── Imputación de pagos desde Excel ───────────────────

@login_required
def imputar_pagos_excel(request):
    if request.method == "POST":
        form = ImputacionExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            df = pd.read_excel(archivo)

            hoy = timezone.now().date()
            imputaciones = 0

            for _, row in df.iterrows():
                try:
                    id_operacion = row["ID Operacion"]
                    saldo = Decimal(str(row["Saldo"]))

                    partes = id_operacion.split("-")
                    if len(partes) != 4:
                        messages.warning(request, f"ID Operación inválido: {id_operacion}")
                        continue

                    nombre_aseg, nombre_ramo, num_poliza, num_endoso = partes

                    plan = PlanPago.objects.get(
                        aseguradora__nombre=nombre_aseg.strip(),
                        ramo__nombre=nombre_ramo.strip(),
                        poliza__numero=num_poliza.strip(),
                        endoso=num_endoso.strip(),
                    )

                    cuotas = plan.cuotas.filter(importe__gt=0).order_by("vencimiento")
                    for cuota in cuotas:
                        if saldo <= 0:
                            break
                        imputado = min(cuota.importe, saldo)
                        saldo -= imputado
                        cuota.importe -= imputado
                        cuota.save()
                        imputaciones += 1

                except PlanPago.DoesNotExist:
                    messages.warning(request, f"No se encontró el plan: {id_operacion}")
                except Exception as e:
                    messages.error(request, f"Error procesando {id_operacion}: {str(e)}")

            if imputaciones:
                messages.success(request, f"Se imputaron pagos en {imputaciones} cuota(s).")
            else:
                messages.warning(request, "No se realizaron imputaciones. Verifique los datos del archivo.")

            return redirect("imputar_pagos_excel")

    else:
        form = ImputacionExcelForm()

    return render(request, "gestion_cobranzas/planes/imputar_excel.html", {"form": form})


# ──────────────────── Exportaciones (Excel / PDF) ───────────────────────

@login_required
def exportar_planes_excel(request):
    poliza_query = request.GET.get("poliza", "")
    planes = PlanPago.objects.all()

    if poliza_query:
        planes = planes.filter(poliza__numero__icontains=poliza_query)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planes de Pago"

    headers = ["Aseguradora", "Ramo", "Póliza", "Endoso", "Moneda", "Cuotas", "Saldo"]
    ws.append(headers)

    for plan in planes:
        cuotas = plan.cuotas.all()
        saldo = sum([cuota.importe for cuota in cuotas])
        ws.append(
            [
                plan.aseguradora.nombre,
                plan.ramo.nombre,
                plan.poliza.numero,
                plan.endoso,
                plan.moneda,
                cuotas.count(),
                f"{plan.moneda} {saldo:,.2f}".replace(",", ".").replace(".", ",", 1),
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=planes_de_pago.xlsx"
    wb.save(response)
    return response


@login_required
def exportar_excel_cuotas(request, plan_id):
    plan = get_object_or_404(PlanPago, id=plan_id)
    cuotas = plan.cuotas.all().order_by("numero")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuotas"

    headers = ["Cuota", "Vencimiento", "Premio", "Saldo", "Estado"]
    ws.append(headers)

    for cuota in cuotas:
        saldo = cuota.importe
        importe_original = cuota.importe_original
        if saldo == 0:
            estado = "Pagada"
        elif saldo < importe_original:
            estado = "Parcial"
        else:
            estado = "Impaga"

        ws.append(
            [
                cuota.numero,
                cuota.vencimiento.strftime("%d/%m/%Y"),
                float(cuota.importe_original),
                float(cuota.importe),
                estado,
            ]
        )

    for col_num, column_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Cuotas_Plan_{plan.poliza.numero}_Endoso_{plan.endoso}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def exportar_pdf_cuotas(request, plan_id):
    plan = get_object_or_404(PlanPago, id=plan_id)
    cuotas = plan.cuotas.all().order_by("numero")

    pagadas = parciales = impagas = 0
    saldo_pendiente = 0

    for cuota in cuotas:
        saldo = cuota.importe
        importe_original = cuota.importe_original or 0
        saldo_pendiente += saldo

        if saldo == 0:
            pagadas += 1
        elif saldo < importe_original:
            parciales += 1
        else:
            impagas += 1

    total_premio = sum(cuota.importe_original or 0 for cuota in cuotas)
    total_saldo = sum(cuota.importe for cuota in cuotas)

    template = get_template("gestion_cobranzas/planes/detalle_plan_pago_pdf.html")
    html = template.render(
        {
            "plan": plan,
            "cuotas": cuotas,
            "pagadas": pagadas,
            "parciales": parciales,
            "impagas": impagas,
            "saldo_pendiente": saldo_pendiente,
            "total_premio": total_premio,
            "total_saldo": total_saldo,
        }
    )

    response = HttpResponse(content_type="application/pdf")
    nombre_archivo = f"Detalle_Plan_Pagos_Poliza_{plan.poliza.numero}_Endoso_{plan.endoso}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

    HTML(string=html).write_pdf(response)
    return response


# ───────────────────────────── Reportes ────────────────────────────────

@login_required
def reportes(request):
    tipo = request.GET.get("tipo")
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")

    resultados = []
    rango_activo = False

    if desde and hasta and tipo:
        try:
            desde_fecha = datetime.strptime(desde, "%Y-%m-%d")
            hasta_fecha = datetime.strptime(hasta, "%Y-%m-%d")

            rango_activo = True

            if tipo == "planes":
                resultados = LogCargaMasiva.objects.filter(
                    fecha_carga__date__range=(desde_fecha, hasta_fecha)
                )
            elif tipo == "imputaciones":
                resultados = LogImputacion.objects.filter(
                    fecha_carga__date__range=(desde_fecha, hasta_fecha)
                )

        except ValueError:
            messages.error(request, "Formato de fechas inválido.")

    return render(
        request,
        "gestion_cobranzas/reportes/reportes.html",
        {
            "resultados": resultados,
            "tipo": tipo,
            "desde": desde,
            "hasta": hasta,
            "rango_activo": rango_activo,
        },
    )
