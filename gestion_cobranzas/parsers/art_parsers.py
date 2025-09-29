import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# ──────────────────────────────────────────────────────────────────────────────
# UTILIDADES COMUNES
# ──────────────────────────────────────────────────────────────────────────────
def _clean_number(value):
    """
    Convierte textos como '12.345,67', '  -2.345,50  ', '$ (1.234,10)' a un
    `float` preservando el signo. Reglas:
      • Elimina símbolo $ y espacios.
      • Si el número viene entre paréntesis ⇒ negativo.
      • Quita puntos de miles y convierte coma decimal a punto.
    """
    if pd.isna(value):
        return 0.0

    txt = str(value).strip()

    # Detectar formato '(123,45)' => negativo
    neg = False
    if txt.startswith('(') and txt.endswith(')'):
        neg = True
        txt = txt[1:-1]

    # Remover signo positivo/negativo explícito y capturarlo
    if txt.startswith('-'):
        neg = True
        txt = txt[1:]
    elif txt.startswith('+'):
        txt = txt[1:]

    # Quitar símbolo de moneda y espacios
    txt = txt.replace('$', '').replace(' ', '')

    # Si tiene coma decimal, quitar puntos de miles y cambiar coma por punto
    if ',' in txt:
        txt = txt.replace('.', '').replace(',', '.')
    # Si sólo tiene puntos como decimal (caso 1234.56) ya está bien

    try:
        num = float(txt)
    except ValueError:
        # Si no se puede convertir, devuelve 0
        return 0.0

    return -num if neg else num


# ──────────────────────────────────────────────────────────────────────────────
# PARSER DE ANDINA
# ──────────────────────────────────────────────────────────────────────────────
def parse_andina(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Lee el Excel de Andina y devuelve un DataFrame normalizado con columnas:
    Periodo | Contrato | Deuda total | Aseguradora
    La Deuda total se calcula NETEANDO (sumando signos + y -) la columna 'Saldo'
    agrupada por 'Numero Poliza'.
    """
    df = pd.read_excel(file_path, dtype=str)

    # Normalizar encabezados
    df.columns = df.columns.str.strip()

    # Limpiar monto, preservar signo
    df['Saldo'] = df['Saldo'].apply(_clean_number)

    # Agrupar: Numero Poliza = contrato, suma neta de saldo
    neto = (
        df.dropna(subset=['Numero Poliza'])
          .groupby('Numero Poliza', as_index=False)['Saldo']
          .sum()
          .rename(columns={'Numero Poliza': 'Contrato',
                           'Saldo': 'Deuda total'})
    )

    # Añadir columnas fijas
    neto.insert(0, 'Periodo', periodo)
    neto['Aseguradora'] = 'Andina'

    return neto[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]


# ──────────────────────────────────────────────────────────────────────────────
# PARSER DE ASOCIART
# ──────────────────────────────────────────────────────────────────────────────
def parse_asociart(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Lee el XLSX de Asociart ignorando celdas combinadas.
    Devuelve: Periodo | Contrato | Deuda total | Aseguradora
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # 1) localizar la fila donde estén las palabras clave
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        lowered = [str(c).lower() if c else "" for c in row]
        if 'contrato' in "".join(lowered) and 'deuda' in "".join(lowered):
            header_row = idx
            headers = [str(c).strip() if c else "" for c in row]
            break

    if header_row is None:
        raise ValueError(f"No se halló cabecera en {file_path.name}")

    # 2) volcar a DataFrame desde la fila siguiente
    data = ws.iter_rows(min_row=header_row + 1, values_only=True)
    df = pd.DataFrame(list(data), columns=headers).dropna(how='all')

    # 3) elegir columnas relevantes (tolerante a mayúsc./tildes/espacios)
    col_contrato = next(c for c in df.columns if 'contrato' in c.lower())
    col_deuda    = next(c for c in df.columns if 'deuda'    in c.lower())

    out = (
        df[[col_contrato, col_deuda]]
          .rename(columns={col_contrato: 'Contrato',
                           col_deuda:   'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Asociart'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE BERKLEY
# ───────────────────────────────────────────────────────────
def parse_berkley(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas clave:
      • Contrato  →  'NRO. CONTRATO'
      • Deuda     →  'SALDO TOTAL'
    Devuelve: Periodo | Contrato | Deuda total | Aseguradora
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    out = (
        df[['NRO. CONTRATO', 'SALDO TOTAL']]
          .rename(columns={'NRO. CONTRATO': 'Contrato',
                           'SALDO TOTAL':  'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Berkley'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE EXPERTA
# ───────────────────────────────────────────────────────────
def parse_experta(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas:
      • Contrato  →  'Numero'
      • Deuda     →  'Saldo Total'   (NEGATIVO → pasar a positivo)
    Devuelve: Periodo | Contrato | Deuda total | Aseguradora
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    out = (
        df[['Numero', 'Saldo Total']]
          .rename(columns={'Numero': 'Contrato',
                           'Saldo Total': 'Deuda total'})
    )
    # limpia separador y convierte a número positivo
    out['Deuda total'] = out['Deuda total'].apply(_clean_number).abs()

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Experta'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE FEDERACIÓN PATRONAL
# ───────────────────────────────────────────────────────────
def parse_fede_patr(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas:
      • Contrato                 → 'Contrato'
      • Cuota adeudada           → 'Cuota adeudada'
      • Interés adeudado         → 'Interés adeudado'
    Deuda total = Cuota adeudada + Interés adeudado
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    # Limpieza numérica y conversión a float
    df['Cuota adeudada']   = df['Cuota adeudada'].apply(_clean_number)
    df['Interés adeudado'] = df['Interés adeudado'].apply(_clean_number)

    df['Deuda total'] = df['Cuota adeudada'] + df['Interés adeudado']

    out = df[['Contrato', 'Deuda total']].copy()
    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Federación Patronal'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE GALENO
# ───────────────────────────────────────────────────────────
def parse_galeno(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas clave en Galeno:
      • Contrato  →  'Poliza'
      • Deuda     →  'Saldo'
    Devuelve: Periodo | Contrato | Deuda total | Aseguradora
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    out = (
        df[['Poliza', 'Saldo']]
          .rename(columns={'Poliza': 'Contrato', 'Saldo': 'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Galeno'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE LA SEGUNDA
# ───────────────────────────────────────────────────────────
def parse_la_segunda(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas clave en La Segunda:
      • Contrato  →  'Contrato'
      • Deuda     →  'Saldo'
    Devuelve: Periodo | Contrato | Deuda total | Aseguradora
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    out = (
        df[['Contrato', 'Saldo']]
          .rename(columns={'Contrato': 'Contrato',   # se deja igual
                           'Saldo':    'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'La Segunda'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE OMINT
# ───────────────────────────────────────────────────────────
def parse_omint(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas clave en Omint:
      • Contrato  →  'Nro. Contrato'
      • Deuda     →  'Saldo Cuenta Corriente'
    Devuelve: Periodo | Contrato | Deuda total | Aseguradora
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    out = (
        df[['Nro. Contrato', 'Saldo Cuenta Corriente']]
          .rename(columns={'Nro. Contrato': 'Contrato',
                           'Saldo Cuenta Corriente': 'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Omint'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE PREVENCIÓN ART
# ───────────────────────────────────────────────────────────
def parse_prevencion(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas clave en Prevención:
      • Contrato      →  'Contrato'
      • Deuda total   →  'Deuda Capital al Último Período Cerrado'
    Devuelve: Periodo | Contrato | Deuda total | Aseguradora
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    # Algunos archivos vienen con el título abreviado; busquemos por palabras
    col_contrato = next(c for c in df.columns if 'contrato' in c.lower())
    col_deuda    = next(
        c for c in df.columns
        if 'deuda' in c.lower() and 'capital' in c.lower()
    )

    out = (
        df[[col_contrato, col_deuda]]
          .rename(columns={col_contrato: 'Contrato',
                           col_deuda:    'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Prevención'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE PROVINCIA ART
# ───────────────────────────────────────────────────────────
def parse_provincia(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas en Provincia:
      • Contrato  →  'CONTRATO'
      • Deuda     →  'SALDO CON INTERESES'
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    out = (
        df[['CONTRATO', 'SALDO CON INTERESES']]
          .rename(columns={'CONTRATO': 'Contrato',
                           'SALDO CON INTERESES': 'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'Provincia'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]

# ───────────────────────────────────────────────────────────
# PARSER DE SMG ART
# ───────────────────────────────────────────────────────────
def parse_smg(file_path: Path, periodo: str) -> pd.DataFrame:
    """
    Columnas en SMG:
      • Contrato  →  'Contrato'
      • Deuda     →  'Premio Saldo Acum.'
    """
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    out = (
        df[['Contrato', 'Premio Saldo Acum.']]
          .rename(columns={'Contrato': 'Contrato',
                           'Premio Saldo Acum.': 'Deuda total'})
    )
    out['Deuda total'] = out['Deuda total'].apply(_clean_number)

    out.insert(0, 'Periodo', periodo)
    out['Aseguradora'] = 'SMG'
    return out[['Periodo', 'Contrato', 'Deuda total', 'Aseguradora']]
